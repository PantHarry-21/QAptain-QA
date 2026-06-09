from __future__ import annotations
import asyncio
import io
import json

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import openpyxl
import csv

from app.db.session import get_db
from app.db.models import (
    User, Scenario, ScenarioPriority, ExecutionPlan, ExecutionRun,
    Environment, Credential, ApplicationModule,
)
from app.core.dependencies import get_current_user
from app.schemas.scenario import (
    ScenarioCreate, ScenarioResponse,
    ExecutionPlanRequest, ExecutionPlanResponse,
    ExecutionTrigger, ExecutionRunResponse,
)
from app.intelligence.scenario_planner import ScenarioPlanner
from app.intelligence.ai_client import get_ai_client
from app.intelligence.exploratory_engine import ExploratoryTestEngine
from app.intelligence.business_rule_engine import BusinessRuleEngine
from app.intelligence.smart_scenario_generator import SmartScenarioGenerator
from app.jobs.execution_job import enqueue_execution, enqueue_batch_execution

router = APIRouter()


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _parse_priority(value) -> ScenarioPriority:
    raw = str(value or "").strip().lower()
    if raw in ("critical", "p1", "blocker", "blocking", "showstopper", "block"):
        return ScenarioPriority.CRITICAL
    if raw in ("high", "p2", "major", "important", "must", "must-have"):
        return ScenarioPriority.HIGH
    if raw in ("low", "p4", "p5", "minor", "trivial", "nice to have", "nice-to-have"):
        return ScenarioPriority.LOW
    return ScenarioPriority.MEDIUM


def _enrich_scenario(s: Scenario, modules_by_id: dict) -> dict:
    """Return a scenario dict with module_name/url fields for the frontend."""
    mod = modules_by_id.get(s.module_id) if s.module_id else None
    return {
        "id": s.id,
        "application_id": s.application_id,
        "title": s.title,
        "description": s.description,
        "priority": s.priority.value if s.priority else "MEDIUM",
        "tags": s.tags or [],
        "module_id": s.module_id,
        "module_name": mod.name if mod else None,
        "module_url": mod.url_pattern if mod else None,
        "source": s.source,
        "is_active": s.is_active,
        "created_at": s.created_at.isoformat() if s.created_at else None,
    }


async def _load_modules_by_id(db: AsyncSession, scenarios: list[Scenario]) -> dict:
    module_ids = {s.module_id for s in scenarios if s.module_id}
    if not module_ids:
        return {}
    result = await db.execute(
        select(ApplicationModule).where(ApplicationModule.id.in_(module_ids))
    )
    return {m.id: m for m in result.scalars().all()}


# ─── Module Auto-Mapping ──────────────────────────────────────────────────────

def _keyword_match_module(title: str, modules: list) -> str | None:
    """
    Match a scenario title to its best-fit module using keyword/substring scoring.
    Returns module_id when confident (score ≥ 0.4), else None.

    Examples that match:
      "Test Add Generic Master"       → module named "Generic Master"   (substring match)
      "Verify CRUD for User Mgmt"     → module named "User Management"  (word overlap)
    """
    title_lower = title.lower()
    best_id: str | None = None
    best_score = 0.0

    for mod in modules:
        mod_name = mod.name.lower()
        # Substring: full module name found inside the scenario title
        if mod_name in title_lower:
            score = len(mod_name) / max(len(title_lower), 1) + 0.5
        else:
            title_words = {w for w in title_lower.split() if len(w) > 2}
            mod_words = {w for w in mod_name.split() if len(w) > 2}
            if not mod_words:
                continue
            overlap = len(title_words & mod_words)
            score = overlap / len(mod_words)

        if score > best_score:
            best_score = score
            best_id = mod.id

    return best_id if best_score >= 0.4 else None


async def _auto_map_scenarios(
    db: AsyncSession,
    application_id: str,
    scenarios: list[Scenario],
) -> int:
    """
    Assign module_id to unmapped scenarios using:
      1. Fast keyword/substring matching against module names
      2. AI bulk mapping for the remainder (one call, all unmapped at once)
    Returns the count of scenarios successfully mapped.
    """
    mods_result = await db.execute(
        select(ApplicationModule).where(ApplicationModule.application_id == application_id)
    )
    modules = mods_result.scalars().all()
    if not modules:
        return 0

    unmapped: list[Scenario] = []
    mapped = 0

    for s in scenarios:
        if s.module_id:
            continue
        mid = _keyword_match_module(s.title, modules)
        if mid:
            s.module_id = mid
            mapped += 1
        else:
            unmapped.append(s)

    # AI fallback — batch all remaining unmapped in a single call
    if unmapped:
        try:
            ai = get_ai_client()
            module_list = [
                {"id": m.id, "name": m.name, "url": m.url_pattern or "",
                 "description": (m.description or "")[:120]}
                for m in modules
            ]
            scenario_list = [
                {"id": s.id, "title": s.title,
                 "description": (s.description or "")[:150]}
                for s in unmapped[:50]
            ]
            resp = await asyncio.wait_for(
                ai.complete(
                    system=(
                        "You are a QA test management assistant. "
                        "Map each test scenario to the single most appropriate module. "
                        'Return ONLY JSON: {"mappings": [{"scenario_id": "...", "module_id": "..."}]}. '
                        "Omit scenarios where no module fits."
                    ),
                    user=(
                        f"MODULES:\n{json.dumps(module_list)}\n\n"
                        f"SCENARIOS TO MAP:\n{json.dumps(scenario_list)}\n\n"
                        "For each scenario choose the module whose name / purpose best matches "
                        "the scenario subject. If the scenario mentions 'Generic Master', map to "
                        "the 'Generic Master' module, etc."
                    ),
                    json_mode=True,
                    fast=True,
                    max_tokens=2000,
                ),
                timeout=30.0,
            )
            valid_ids = {m.id for m in modules}
            mapping: dict[str, str] = {
                m["scenario_id"]: m["module_id"]
                for m in resp.json().get("mappings", [])
                if m.get("scenario_id") and m.get("module_id") in valid_ids
            }
            for s in unmapped:
                mid = mapping.get(s.id)
                if mid:
                    s.module_id = mid
                    mapped += 1
        except Exception:
            pass  # mapping is best-effort; don't fail the import

    if mapped:
        await db.commit()
    return mapped


def _extract_docx_text(content: bytes) -> str:
    from docx import Document  # python-docx
    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_pdf_text(content: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_excel_test_cases(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
    headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]

    # ── Column detection ──────────────────────────────────────────────────────
    # Ordered preference lists: earlier match wins.
    # ID/code columns are listed last so name/title columns win when both exist.
    TITLE_PREFERENCES = [
        # Tier 1 — explicit name/title columns
        {"title", "test case name", "test case title", "tc name", "tc title",
         "case name", "scenario name", "scenario title", "test name", "name"},
        # Tier 2 — generic scenario/case without "id"
        {"scenario", "test case", "test_case", "testcase", "test scenario", "case"},
        # Tier 3 — bare "test" or any id-like column (last resort)
        {"test", "test case id", "tc id", "test id", "tc no", "test no"},
    ]
    DESC_EXACT = {
        "description", "steps", "test steps", "test description", "desc",
        "expected result", "expected output", "expected", "details",
        "test steps/actions", "steps/actions", "test steps & expected results",
        "objective", "test objective", "preconditions", "pre-conditions",
    }
    PRIORITY_EXACT = {"priority", "severity", "criticality", "importance", "level"}

    def _find_title_col() -> int | None:
        for tier in TITLE_PREFERENCES:
            for i, h in enumerate(headers):
                if h in tier:
                    return i
        # Substring fallback: prefer "title" > "name" > "scenario" > "case"
        for kw in ("title", "name", "scenario", "case"):
            for i, h in enumerate(headers):
                if h and kw in h and "id" not in h:
                    return i
        # Absolute last resort: first non-empty header
        return next((i for i, h in enumerate(headers) if h), None)

    def _find_col(exact_set: set, substr_keywords: tuple) -> int | None:
        for i, h in enumerate(headers):
            if h in exact_set:
                return i
        for i, h in enumerate(headers):
            if h and any(kw in h for kw in substr_keywords):
                return i
        return None

    title_col = _find_title_col()
    desc_col  = _find_col(DESC_EXACT,    ("description", "steps", "expected", "detail", "action", "objective"))
    prio_col  = _find_col(PRIORITY_EXACT, ("priority", "severity"))

    if title_col is None:
        return []

    # ── Row extraction ────────────────────────────────────────────────────────
    test_cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        if all(v is None or str(v).strip() == "" for v in cells):
            continue

        raw_title = cells[title_col] if title_col < len(cells) else None
        title = str(raw_title).strip() if raw_title is not None else ""
        if not title or title.lower() in ("none", "nan", "-", "n/a", ""):
            continue

        description = ""
        if desc_col is not None and desc_col < len(cells):
            dv = cells[desc_col]
            description = str(dv).strip() if dv is not None else ""

        priority = "MEDIUM"
        if prio_col is not None and prio_col < len(cells):
            pv = cells[prio_col]
            if pv is not None:
                priority = str(pv).strip()

        test_cases.append({"title": title, "description": description, "priority": priority})

    return test_cases


_DOC_EXTRACT_SYSTEM = """You are a test case extraction engine.

Given document text from a test cases file, extract ALL test cases as a JSON array.

Return ONLY valid JSON — no markdown, no explanation:
{
  "test_cases": [
    {
      "title": "Short imperative title starting with a verb (Verify / Create / Test / Validate / Ensure)",
      "description": "What to test — 1–2 sentences describing the test objective and expected outcome",
      "priority": "CRITICAL|HIGH|MEDIUM|LOW"
    }
  ]
}

Rules:
- Extract EVERY distinct test case, test step, or numbered item from the document
- If the document has numbered items like "1. Test login", extract each one
- Infer priority: "critical / must / block / blocker" → CRITICAL; "should / important / must-have" → HIGH; default → MEDIUM
- Keep titles concise (under 100 chars) and actionable
- If the document is a general description with no clear test cases, infer sensible test cases from the described functionality"""


async def _ai_extract_test_cases(module_name: str, module_url: str, text: str) -> list[dict]:
    """Call AI to extract test cases from free-form document text (DOCX/PDF)."""
    ai = get_ai_client()
    response = await ai.complete(
        system=_DOC_EXTRACT_SYSTEM,
        user=f"MODULE: {module_name}\nMODULE URL: {module_url}\n\nDOCUMENT:\n{text[:8000]}",
        json_mode=True,
        max_tokens=4000,
    )
    try:
        extracted = response.json()
        return extracted.get("test_cases", [])
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"AI failed to parse test cases from the document: {exc}",
        ) from exc


# ─── List ─────────────────────────────────────────────────────────────────────

@router.get("")
async def list_scenarios(
    application_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Scenario)
        .where(Scenario.application_id == application_id, Scenario.is_active == True)
        .order_by(Scenario.created_at.desc())
    )
    scenarios = result.scalars().all()
    modules_by_id = await _load_modules_by_id(db, scenarios)
    return [_enrich_scenario(s, modules_by_id) for s in scenarios]


# ─── Create ───────────────────────────────────────────────────────────────────

@router.post("", status_code=201)
async def create_scenario(
    payload: ScenarioCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    scenario = Scenario(
        application_id=payload.application_id,
        title=payload.title,
        description=payload.description,
        priority=payload.priority,
        tags=payload.tags,
        module_id=payload.module_id,
        source="manual",
        created_by=current_user.id,
    )
    db.add(scenario)
    await db.commit()

    # Auto-map to a knowledge graph module if no explicit module was provided
    if not payload.module_id:
        await _auto_map_scenarios(db, payload.application_id, [scenario])

    modules_by_id = await _load_modules_by_id(db, [scenario])
    return _enrich_scenario(scenario, modules_by_id)


# ─── Update ───────────────────────────────────────────────────────────────────

class ScenarioUpdate(BaseModel):
    title: str | None = None
    description: str | None = None
    priority: str | None = None
    tags: list[str] | None = None
    module_id: str | None = None  # set to assign; omit to keep current

# ─── Bulk delete by module (must be before /{scenario_id} to avoid route shadowing) ─

@router.delete("/bulk/by-module", status_code=200)
async def delete_scenarios_by_module(
    application_id: str,
    module_id: str | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Soft-delete all scenarios for a module (or all unassigned if module_id=none)."""
    query = select(Scenario).where(
        Scenario.application_id == application_id,
        Scenario.is_active == True,
    )
    if module_id and module_id != "__none__":
        query = query.where(Scenario.module_id == module_id)
    else:
        query = query.where(Scenario.module_id == None)  # noqa: E711

    result = await db.execute(query)
    scenarios = result.scalars().all()
    for s in scenarios:
        s.is_active = False
    await db.commit()
    return {"deleted": len(scenarios)}


# ─── Update ───────────────────────────────────────────────────────────────────

@router.put("/{scenario_id}")
async def update_scenario(
    scenario_id: str,
    payload: ScenarioUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Scenario).where(Scenario.id == scenario_id))
    scenario = result.scalar_one_or_none()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
    if payload.title is not None:
        scenario.title = payload.title[:512]
    if payload.description is not None:
        scenario.description = payload.description
    if payload.priority is not None:
        scenario.priority = _parse_priority(payload.priority)
    if payload.tags is not None:
        scenario.tags = payload.tags
    # module_id: explicit value assigns; "module_id" key not in request → keep current
    if "module_id" in payload.model_fields_set:
        scenario.module_id = payload.module_id or None
    await db.commit()
    await db.refresh(scenario)
    modules_by_id = await _load_modules_by_id(db, [scenario])
    return _enrich_scenario(scenario, modules_by_id)


# ─── Delete (soft) ────────────────────────────────────────────────────────────

@router.delete("/{scenario_id}", status_code=204)
async def delete_scenario(
    scenario_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Scenario).where(Scenario.id == scenario_id))
    scenario = result.scalar_one_or_none()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
    scenario.is_active = False
    await db.commit()


# ─── Auto-map to Knowledge Graph modules ─────────────────────────────────────

@router.post("/auto-map-modules")
async def auto_map_modules(
    application_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Map all unlinked scenarios to their knowledge graph module using
    keyword matching + AI fallback.  Call this after any bulk import or
    whenever scenarios show 'No Module' in the UI.
    """
    result = await db.execute(
        select(Scenario).where(
            Scenario.application_id == application_id,
            Scenario.is_active == True,
            Scenario.module_id == None,  # noqa: E711
        )
    )
    unmapped = result.scalars().all()
    if not unmapped:
        return {"mapped": 0, "total_unmapped": 0,
                "message": "All scenarios already have modules assigned"}

    mapped = await _auto_map_scenarios(db, application_id, list(unmapped))
    return {
        "mapped": mapped,
        "total_unmapped": len(unmapped),
        "message": f"Mapped {mapped} of {len(unmapped)} scenarios to knowledge graph modules",
    }


# ─── Import: Excel ────────────────────────────────────────────────────────────

@router.post("/import/excel")
async def import_from_excel(
    application_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    content = await file.read()
    test_cases = _extract_excel_test_cases(content)
    if not test_cases:
        raise HTTPException(status_code=422, detail="No test cases found. Check that your file has Title/Description columns.")

    created: list[Scenario] = []
    for tc in test_cases:
        title = tc.get("title", "").strip()
        if not title:
            continue
        scenario = Scenario(
            application_id=application_id,
            title=title[:512],
            description=tc.get("description", ""),
            priority=_parse_priority(tc.get("priority", "MEDIUM")),
            tags=[t.strip() for t in str(tc.get("tags", "")).split(",") if t.strip()],
            source="excel",
            created_by=current_user.id,
        )
        db.add(scenario)
        created.append(scenario)

    await db.commit()

    # Auto-map every scenario to its knowledge graph module
    mapped = await _auto_map_scenarios(db, application_id, created)

    return {
        "imported": len(created),
        "mapped_to_modules": mapped,
        "titles": [s.title for s in created[:20]],
    }


# ─── Import: CSV ──────────────────────────────────────────────────────────────

@router.post("/import/csv")
async def import_from_csv(
    application_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    created: list[Scenario] = []
    for row in reader:
        title = row.get("title") or row.get("scenario") or row.get("test_case")
        if not title:
            continue
        scenario = Scenario(
            application_id=application_id,
            title=str(title).strip(),
            description=row.get("description", ""),
            priority=_parse_priority(row.get("priority")),
            tags=[t.strip() for t in (row.get("tags", "") or "").split(",") if t.strip()],
            source="csv",
            created_by=current_user.id,
        )
        db.add(scenario)
        created.append(scenario)

    await db.commit()

    # Auto-map every scenario to its knowledge graph module
    mapped = await _auto_map_scenarios(db, application_id, created)

    return {
        "imported": len(created),
        "mapped_to_modules": mapped,
        "titles": [s.title for s in created[:10]],
    }


# ─── Import: Document (DOCX / PDF) ────────────────────────────────────────────

@router.post("/import/document")
async def import_from_document(
    application_id: str,
    module_name: str,
    module_url: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload a DOCX or PDF test-cases file.
    AI extracts every test case. All are saved as Scenarios linked to a module
    with the given URL — so execution navigates there automatically.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith((".xlsx", ".xls")):
        test_cases = _extract_excel_test_cases(content)
        if not test_cases:
            raise HTTPException(
                status_code=422,
                detail=(
                    "No test cases found in the Excel file. "
                    "Ensure the first row contains column headers and at least one column "
                    "is named: Title, Name, Scenario, Test Case, Test Case Name, or similar."
                ),
            )
    elif filename.endswith(".docx"):
        raw_text = _extract_docx_text(content)
        if not raw_text.strip():
            raise HTTPException(status_code=400, detail="Could not extract text from the document")
        test_cases = await _ai_extract_test_cases(module_name, module_url, raw_text)
    elif filename.endswith(".pdf"):
        raw_text = _extract_pdf_text(content)
        if not raw_text.strip():
            raise HTTPException(status_code=400, detail="Could not extract text from the document")
        test_cases = await _ai_extract_test_cases(module_name, module_url, raw_text)
    else:
        raise HTTPException(
            status_code=400,
            detail="Only .docx, .pdf, .xlsx, and .xls files are supported",
        )

    if not test_cases:
        raise HTTPException(status_code=422, detail="No test cases found in the document")

    # Find or create module
    mod_result = await db.execute(
        select(ApplicationModule).where(
            ApplicationModule.application_id == application_id,
            ApplicationModule.name == module_name,
        ).limit(1)
    )
    module = mod_result.scalar_one_or_none()
    if not module:
        module = ApplicationModule(
            application_id=application_id,
            name=module_name,
            url_pattern=module_url,
            description="Imported via document upload",
        )
        db.add(module)
        await db.flush()
    else:
        module.url_pattern = module_url

    # Create scenarios
    created = []
    for tc in test_cases:
        title = (tc.get("title") or "").strip()
        if not title:
            continue
        scenario = Scenario(
            application_id=application_id,
            module_id=module.id,
            title=title,
            description=tc.get("description", ""),
            priority=_parse_priority(tc.get("priority", "MEDIUM")),
            tags=["document-import"],
            source="document",
            created_by=current_user.id,
        )
        db.add(scenario)
        created.append(title)

    await db.commit()
    return {
        "imported": len(created),
        "module": module_name,
        "module_url": module_url,
        "module_id": module.id,
        "titles": created[:20],
    }


# ─── Run Batch ────────────────────────────────────────────────────────────────

class RunBatchPayload(BaseModel):
    scenario_ids: list[str]
    execution_mode: str = "functional"
    environment_id: str


@router.post("/run-batch")
async def run_batch(
    payload: RunBatchPayload,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Execute multiple scenarios with a single browser session (BeforeAll pattern).
    Optimized: bulk-fetches all scenarios, builds all plans in one DB commit,
    then enqueues execution — returns run IDs in ~100ms regardless of batch size.
    """
    import uuid as _uuid
    import traceback as _tb
    import structlog as _sl
    _log = _sl.get_logger()

    if not payload.scenario_ids:
        raise HTTPException(status_code=400, detail="No scenario IDs provided")

    cap = 50
    requested_ids = list(dict.fromkeys(payload.scenario_ids[:cap]))  # dedup, preserve order

    # ── 1. Bulk-fetch all scenarios in ONE query ──────────────────────────────
    result = await db.execute(
        select(Scenario).where(Scenario.id.in_(requested_ids))
    )
    scenario_lookup: dict[str, Scenario] = {s.id: s for s in result.scalars().all()}

    errors: list[dict] = [
        {"scenario_id": sid, "error": "Scenario not found"}
        for sid in requested_ids if sid not in scenario_lookup
    ]

    # ── 2. Build ALL plan objects (no AI, no commit) in a single pass ─────────
    planner = ScenarioPlanner(db)
    plan_entries: list[tuple] = []  # (plan_obj, scenario_title)
    for sid in requested_ids:
        if sid not in scenario_lookup:
            continue
        scenario = scenario_lookup[sid]
        plan_data = planner._fallback_plan(scenario)
        from app.db.models import ExecutionPlan
        from config import settings as _cfg
        plan = ExecutionPlan(
            scenario_id=scenario.id,
            execution_mode=payload.execution_mode,
            plan_data=plan_data,
            ai_reasoning="Fallback — AI reasoning deferred to execution time",
            semantic_intent={},
            workflow_stages=[],
            risk_score=5,
            estimated_duration_seconds=len(plan_data.get("steps", [])) * 5,
            created_by_model="fallback",
        )
        db.add(plan)
        plan_entries.append((plan, scenario.title))

    if not plan_entries:
        return {"runs": errors, "total": 0, "batch_mode": True}

    # ── 3. ONE commit for all plans ───────────────────────────────────────────
    await db.commit()

    plans = [p for p, _ in plan_entries]
    plan_title_map: dict[str, str] = {p.id: title for p, title in plan_entries}

    # ── 4. Enqueue batch — creates runs + submits thread job ──────────────────
    batch_id = str(_uuid.uuid4())
    try:
        runs = await enqueue_batch_execution(
            db=db,
            plans=plans,
            environment_id=payload.environment_id,
            credential_id=None,
            triggered_by=current_user.id,
            batch_id=batch_id,
        )
        run_summaries = [
            {
                "scenario_id": run.scenario_id,
                "run_id": run.id,
                "title": plan_title_map.get(run.plan_id, ""),
            }
            for run in runs
        ]
        _log.info("Batch enqueued", count=len(runs), batch_id=batch_id)
        return {
            "runs": run_summaries + errors,
            "total": len(run_summaries),
            "batch_id": batch_id,
            "batch_mode": True,
        }
    except Exception as e:
        err_detail = f"Batch enqueue failed: {e}\n{_tb.format_exc()[-300:]}"
        _log.error("Batch enqueue error", error=err_detail)
        raise HTTPException(status_code=500, detail=err_detail)


# ─── Generate plan ────────────────────────────────────────────────────────────

@router.post("/{scenario_id}/plan", response_model=ExecutionPlanResponse)
async def generate_execution_plan(
    scenario_id: str,
    payload: ExecutionPlanRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Scenario).where(Scenario.id == scenario_id))
    scenario = result.scalar_one_or_none()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")

    if not payload.force_regenerate:
        existing = await db.execute(
            select(ExecutionPlan)
            .where(ExecutionPlan.scenario_id == scenario_id)
            .order_by(ExecutionPlan.created_at.desc())
            .limit(1)
        )
        existing_plan = existing.scalar_one_or_none()
        # Only reuse AI-generated plans. Capability-engine and fallback plans are
        # always regenerated — the engines improve over time and cached plans may
        # have stale entity names, bad selectors, etc.
        if existing_plan and existing_plan.created_by_model not in ("fallback", "capability_engine"):
            return ExecutionPlanResponse.model_validate(existing_plan)

    planner = ScenarioPlanner(db)
    try:
        plan = await planner.generate_plan(scenario, payload.execution_mode)
    except Exception as e:
        import structlog
        log = structlog.get_logger()
        log.error("Plan generation failed — using fallback plan", error=str(e))
        plan = await planner.generate_fallback_plan(scenario, payload.execution_mode)
    return ExecutionPlanResponse.model_validate(plan)


# ─── Trigger execution ────────────────────────────────────────────────────────

@router.post("/{scenario_id}/execute", response_model=ExecutionRunResponse, status_code=201)
async def trigger_execution(
    scenario_id: str,
    payload: ExecutionTrigger,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(ExecutionPlan).where(ExecutionPlan.id == payload.plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Execution plan not found")

    run = await enqueue_execution(
        db=db,
        plan=plan,
        environment_id=payload.environment_id,
        credential_id=payload.credential_id,
        triggered_by=current_user.id,
    )
    return ExecutionRunResponse.model_validate(run)


# ─── List runs ────────────────────────────────────────────────────────────────

@router.get("/{scenario_id}/runs", response_model=list[ExecutionRunResponse])
async def list_runs(
    scenario_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ExecutionRun)
        .where(ExecutionRun.scenario_id == scenario_id)
        .order_by(ExecutionRun.created_at.desc())
    )
    return [ExecutionRunResponse.model_validate(r) for r in result.scalars().all()]


# ─── Feature 1: Exploratory Testing ──────────────────────────────────────────

class ExploratoryRequest(BaseModel):
    target: str                  # e.g. "Test Add Product"
    application_id: str


@router.post("/exploratory", status_code=201)
async def generate_exploratory_tests(
    payload: ExploratoryRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Exploratory Testing Engine.

    Given a natural-language target ("Test Add Product"), generates ~18-24
    exploratory test scenarios across 6 categories:
      empty_form, invalid_values, boundary_values, duplicate_data,
      unauthorized_access, max_length

    The AI reads the application's knowledge graph (forms, field validations,
    workflows) to make the tests specific to the actual UI.
    """
    engine = ExploratoryTestEngine(db, get_ai_client())
    scenarios = await engine.generate(
        target=payload.target,
        application_id=payload.application_id,
        user_id=current_user.id,
    )
    modules_by_id = await _load_modules_by_id(db, scenarios)
    return {
        "generated": len(scenarios),
        "target": payload.target,
        "categories": [
            "empty_form", "invalid_values", "boundary_values",
            "duplicate_data", "unauthorized_access", "max_length",
        ],
        "scenarios": [_enrich_scenario(s, modules_by_id) for s in scenarios],
    }


# ─── Feature 2: Business Rule Discovery ──────────────────────────────────────

@router.post("/business-rules/{application_id}", status_code=201)
async def discover_and_generate_business_rule_tests(
    application_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Business Rule Discovery Engine.

    Analyses the application's knowledge graph (form validations, workflow
    preconditions, error paths) to infer implicit business rules, then generates
    a positive test + a negative test for each rule.

    Example inferred rules:
      - Price must be greater than zero
      - Email must be unique
      - Location must be selected before creating a Sample
      - End date cannot be before start date
    """
    engine = BusinessRuleEngine(db, get_ai_client())
    rules, scenarios = await engine.generate_scenarios(
        application_id=application_id,
        user_id=current_user.id,
    )
    modules_by_id = await _load_modules_by_id(db, scenarios)
    return {
        "rules_discovered": len(rules),
        "scenarios_generated": len(scenarios),
        "rules": [
            {
                "id": r.get("id"),
                "name": r.get("name"),
                "category": r.get("category"),
                "description": r.get("description"),
                "entity": r.get("entity"),
                "field": r.get("field"),
                "confidence": r.get("confidence"),
            }
            for r in rules
        ],
        "scenarios": [_enrich_scenario(s, modules_by_id) for s in scenarios],
    }


# ─── Feature 3: Smart Test Generation ────────────────────────────────────────

class SmartGenerateRequest(BaseModel):
    source_type: str    # user_story | requirement | screenshot | workflow | production_logs
    content: str        # text content, or workflow name for source_type=workflow
    application_id: str


@router.post("/generate", status_code=201)
async def smart_generate_scenarios(
    payload: SmartGenerateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Smart Test Generation Engine.

    Converts multiple artifact types into complete test suites
    (happy path + edge cases + negative + regression):

      user_story      → test cases per acceptance criterion
      requirement     → test suite per requirement statement
      screenshot      → tests for all visible interactive elements
      workflow        → tests per workflow stage + precondition violations
      production_logs → regression tests per error pattern
    """
    valid_types = {"user_story", "requirement", "screenshot", "workflow", "production_logs"}
    if payload.source_type not in valid_types:
        raise HTTPException(
            status_code=400,
            detail=f"source_type must be one of: {', '.join(sorted(valid_types))}",
        )

    generator = SmartScenarioGenerator(db, get_ai_client())
    scenarios = await generator.generate(
        source_type=payload.source_type,  # type: ignore[arg-type]
        content=payload.content,
        application_id=payload.application_id,
        user_id=current_user.id,
    )
    modules_by_id = await _load_modules_by_id(db, scenarios)
    return {
        "generated": len(scenarios),
        "source_type": payload.source_type,
        "categories": ["happy_path", "edge_case", "negative", "regression"],
        "scenarios": [_enrich_scenario(s, modules_by_id) for s in scenarios],
    }


# ─── AI Copilot: generate scenarios / user stories from plain description ─────

class AICopilotRequest(BaseModel):
    description: str
    application_id: str
    output_type: str = "scenarios"   # "scenarios" | "user_stories"


@router.post("/ai-copilot/generate")
async def ai_copilot_generate(
    payload: AICopilotRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    AI Copilot: generate test scenarios / user stories from plain-language description.

    When exploration data exists for the application, the AI receives:
      - Exact module name + URL
      - Exact form field labels (Name, Code, Status…)
      - Exact trigger button labels ("New Generic Master")
      - Table structure for list-verification steps
      - Discovered workflow steps from real navigation

    Without exploration data, falls back to generic AI generation.
    """
    import asyncio
    from app.db.models import Application
    from app.intelligence.copilot_context_builder import CopilotContextBuilder

    app_result = await db.execute(select(Application).where(Application.id == payload.application_id))
    app = app_result.scalar_one_or_none()
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # ── Build exploration context ──────────────────────────────────────────────
    ctx = await CopilotContextBuilder(db).build(payload.description, payload.application_id)
    is_user_stories = payload.output_type == "user_stories"

    if ctx["has_exploration_data"]:
        system_prompt, user_prompt = _build_contextual_prompt(app, ctx, payload.description, is_user_stories)
    else:
        system_prompt, user_prompt = _build_generic_prompt(app, ctx, payload.description, is_user_stories)

    ai = get_ai_client()
    try:
        response = await asyncio.wait_for(
            ai.complete(system=system_prompt, user=user_prompt, fast=False, json_mode=True, max_tokens=16000),
            timeout=60.0,
        )
        result = response.json()
        items = result.get("items", [])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI generation failed: {str(e)[:200]}")

    return {
        "output_type": payload.output_type,
        "items": items,
        "application_id": payload.application_id,
        "context_used": ctx["has_exploration_data"],
        "matched_module": ctx.get("module", {}).get("name") if ctx["has_exploration_data"] else None,
    }


# ─── Prompt builders ──────────────────────────────────────────────────────────

def _build_contextual_prompt(
    app,
    ctx: dict,
    description: str,
    is_user_stories: bool,
) -> tuple[str, str]:
    """
    Build a prompt grounded in real exploration data.
    Includes exact module URL, button labels, form fields, and table columns.
    """
    module = ctx["module"]
    operation = ctx["operation"]
    forms = ctx["forms"]
    tables = ctx["tables"]
    elements = ctx["elements"]
    matched_wf = ctx.get("matched_workflow")

    # ── Module section ─────────────────────────────────────────────────────────
    mod_section = f"MODULE: {module['name']}\nURL: {module['url'] or 'not captured'}\n"
    if module.get("description"):
        mod_section += f"Description: {module['description']}\n"

    # ── Form fields section ────────────────────────────────────────────────────
    form_section = ""
    if forms:
        form_section = "\nFORMS (discovered during exploration):\n"
        for form in forms[:3]:
            form_section += f"  Form: {form.get('name', 'Form')}\n"
            if form.get("purpose"):
                form_section += f"  Purpose: {form['purpose']}\n"
            fields = form.get("fields") or []
            if fields:
                form_section += "  Fields:\n"
                for f in fields[:20]:
                    req = " (REQUIRED)" if f.get("required") else ""
                    opts = ""
                    if f.get("options"):
                        opts = f" [options: {', '.join(str(o) for o in f['options'][:5])}]"
                    val = f" — validation: {f['validation']}" if f.get("validation") else ""
                    form_section += f"    - {f['label']}: {f.get('type', 'text')}{req}{opts}{val}\n"
            if form.get("success_message"):
                form_section += f"  Success indicator: {form['success_message']}\n"
            if form.get("submit_action"):
                form_section += f"  Submit action: {form['submit_action']}\n"

    # ── Buttons / elements section ─────────────────────────────────────────────
    button_labels = [e["label"] for e in elements if e["type"] in ("button", "link", "element")]
    field_labels  = [e["label"] for e in elements if e["type"] in ("textbox", "input", "dropdown", "select")]
    elem_section = ""
    if button_labels:
        elem_section += f"\nACTION BUTTONS (exact labels from exploration):\n"
        for b in button_labels[:15]:
            elem_section += f"  - \"{b}\"\n"
    if field_labels:
        elem_section += f"\nFORM INPUTS (exact labels from exploration):\n"
        for f in field_labels[:15]:
            elem_section += f"  - \"{f}\"\n"

    # ── Tables section ─────────────────────────────────────────────────────────
    table_section = ""
    if tables:
        table_section = "\nLIST/TABLE (for verification steps):\n"
        for tbl in tables[:2]:
            table_section += f"  Table: {tbl.get('name', 'List')}\n"
            cols = tbl.get("columns") or []
            if cols:
                table_section += f"  Columns: {', '.join(c.get('name', '') for c in cols[:8])}\n"
            row_actions = tbl.get("row_actions") or []
            if row_actions:
                table_section += f"  Row actions: {', '.join(str(a) for a in row_actions[:5])}\n"
            if tbl.get("has_search"):
                table_section += "  Has search bar: YES\n"

    # ── Workflow section ───────────────────────────────────────────────────────
    wf_section = ""
    if matched_wf:
        wf_section = f"\nDISCOVERED WORKFLOW: {matched_wf['name']} (type: {matched_wf['type']})\n"
        if matched_wf.get("trigger"):
            wf_section += f"  Entry trigger: \"{matched_wf['trigger']}\"\n"
        if matched_wf.get("preconditions"):
            wf_section += f"  Preconditions: {', '.join(matched_wf['preconditions'][:3])}\n"
        stages = matched_wf.get("stages") or []
        if stages:
            wf_section += "  Steps discovered:\n"
            for stage in stages[:8]:
                action = stage.get("action", "")
                expected = stage.get("expected_result", "")
                wf_section += f"    {stage.get('step', '')}. {action}"
                if expected:
                    wf_section += f" → {expected}"
                wf_section += "\n"
        if matched_wf.get("success_indicators"):
            wf_section += f"  Success indicators: {', '.join(str(s) for s in matched_wf['success_indicators'][:3])}\n"

    # ── Operation hint ─────────────────────────────────────────────────────────
    op_hints = {
        "crud_create": "Generate: validation tests (empty, spaces, boundaries) + happy path + list verification + duplicates",
        "crud_update": "Generate: validation on edit form + happy path update + verify changes persist + cancel flow",
        "crud_delete": "Generate: delete confirmation flow + verify removed from list + undo/soft-delete if applicable",
        "crud_read":   "Generate: list loads correctly + pagination + sorting + empty state + data accuracy",
        "search":      "Generate: exact match + partial match + no results + special characters + filter combinations",
    }
    op_instruction = op_hints.get(operation, "Generate comprehensive test scenarios covering happy path and failure cases.")

    ground_truth = f"{mod_section}{form_section}{elem_section}{table_section}{wf_section}"

    if is_user_stories:
        system_prompt = f"""You are a QA Business Analyst. Generate User Stories grounded in real application data.

GROUND TRUTH FROM APPLICATION EXPLORATION:
{ground_truth}

Rules:
- Use EXACT field names, button labels, and module names from the exploration data above
- Reference the actual URL when describing navigation
- acceptance_criteria must be testable Given/When/Then statements using exact UI labels
- test_hints must reference specific fields and buttons by exact name

Return ONLY valid JSON:
{{
  "items": [
    {{
      "title": "As a user, I want to [action using exact module/entity name] so that [benefit]",
      "acceptance_criteria": [
        "Given I navigate to [exact URL], When I click '[exact button]', Then the form opens",
        "Given the form is open, When I fill '[exact field]' and click '[submit button]', Then [success message]",
        "Given the form is open, When I leave '[required field]' empty and click Save, Then a validation error appears"
      ],
      "priority": "HIGH | MEDIUM | LOW",
      "test_hints": [
        "Test '[exact button]' without filling '[required field]'",
        "Test with valid data in all fields",
        "Test search in list for the created record"
      ]
    }}
  ]
}}"""

        user_prompt = f"""Application: {app.name} (base URL: {app.base_url})
User request: {description}
Operation type detected: {operation}
{op_instruction}

Generate 3-5 User Stories using ONLY the exact element names and URLs from the ground truth above."""

    else:
        system_prompt = f"""You are a Senior QA Engineer generating executable test scenarios from real application data.

GROUND TRUTH FROM APPLICATION EXPLORATION:
{ground_truth}

Rules:
- Use EXACT button labels (e.g. "New Generic Master", not "Add button")
- Use EXACT field labels (e.g. "Name", "Code", "Status" — as listed in FORM INPUTS above)
- Include the exact module URL in navigation steps
- Each scenario description must be a step-by-step instruction a Selenium test runner can follow
- For CRUD_CREATE: always include validation test (empty save), spaces test, valid data test, and list verification
- description format: "Navigate to [URL]. Click '[Button]'. [Fill/Assert/Verify steps with exact labels]."

Return ONLY valid JSON:
{{
  "items": [
    {{
      "title": "Verify [specific action] on [exact module name]",
      "description": "Navigate to [exact URL]. Click '[exact button label]'. [Step-by-step with exact field names]. Verify [exact success indicator].",
      "priority": "HIGH | MEDIUM | LOW",
      "category": "happy_path | negative | edge_case | validation | regression"
    }}
  ]
}}"""

        user_prompt = f"""Application: {app.name} (base URL: {app.base_url})
User request: {description}
Operation type detected: {operation}
{op_instruction}

Generate 5-7 test scenarios using ONLY the exact element names, field labels, and URLs from the ground truth above.
Do NOT invent field names or button labels — use what is listed."""

    return system_prompt, user_prompt


def _build_generic_prompt(
    app,
    ctx: dict,
    description: str,
    is_user_stories: bool,
) -> tuple[str, str]:
    """
    Fallback prompt when no exploration data is available.
    Identical to the original behavior.
    """
    all_module_names = ctx.get("all_module_names", [])
    module_hint = ""
    if all_module_names:
        module_hint = "Known modules: " + ", ".join(all_module_names[:10]) + "\n"

    if is_user_stories:
        system_prompt = """You are a QA Business Analyst. Convert a plain description into well-structured User Stories for testing.

Format each user story as:
- title: "As a [role], I want to [action] so that [benefit]"
- acceptance_criteria: list of 3-5 testable conditions (Given/When/Then format)
- priority: HIGH | MEDIUM | LOW
- test_hints: 2-3 specific test scenarios this story implies

Return ONLY valid JSON:
{
  "items": [
    {
      "title": "As an admin, I want to add a product so that inventory is updated",
      "acceptance_criteria": [
        "Given I am on the Products page, When I click Add, Then the Add Product form opens",
        "Given the form is open, When I fill all required fields and click Save, Then the product appears in the list",
        "Given the form is open, When I click Save without filling Name, Then an error is shown"
      ],
      "priority": "HIGH",
      "test_hints": ["Test empty form submission", "Test with valid data", "Test duplicate name"]
    }
  ]
}"""
    else:
        system_prompt = """You are a Senior QA Engineer. Convert a plain description into specific, executable test scenarios.

Include: happy path, negative tests, edge cases, and business rule validations.

Return ONLY valid JSON:
{
  "items": [
    {
      "title": "Verify successful creation with all required fields",
      "description": "Navigate to the module. Click Add/New. Fill all required fields with valid data. Save and verify the record appears in the list.",
      "priority": "HIGH",
      "category": "happy_path"
    }
  ]
}"""

    user_prompt = (
        f"Application: {app.name}\nBase URL: {app.base_url}\n"
        f"{module_hint}"
        f"\nNote: This application has not been explored yet — no element data available. "
        f"Generate best-effort scenarios.\n"
        f"\nUser request: {description}"
    )

    return system_prompt, user_prompt
